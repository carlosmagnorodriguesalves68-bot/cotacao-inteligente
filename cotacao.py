import io
import pandas as pd
import streamlit as st
from openpyxl import load_workbook

# =========================================================
# CONFIGURAÇÃO DA PÁGINA
# =========================================================
st.set_page_config(page_title="CotaBot", layout="wide")

# TOPO COM LOGO + NOME
topo1, topo2 = st.columns([1, 5])

with topo1:
    try:
        st.image("logo.png", width=180)
    except Exception:
        pass

with topo2:
    st.markdown("<h1 style='margin-bottom:0;'>CotaBot</h1>", unsafe_allow_html=True)
    st.markdown(
        "<p style='color:gray; margin-top:0;'>Sua cotação pronta em segundos</p>",
        unsafe_allow_html=True
    )

# =========================================================
# FUNÇÕES AUXILIARES
# =========================================================
def normalizar_texto(texto):
    texto = str(texto).strip().lower()
    trocas = {
        "ç": "c",
        "ã": "a", "á": "a", "à": "a", "â": "a",
        "é": "e", "ê": "e",
        "í": "i",
        "ó": "o", "ô": "o", "õ": "o",
        "ú": "u",
    }
    for k, v in trocas.items():
        texto = texto.replace(k, v)
    return texto


def limpar_ean(serie):
    return (
        serie.astype(str)
        .str.strip()
        .str.replace(".0", "", regex=False)
        .str.replace(" ", "", regex=False)
        .str.replace("-", "", regex=False)
    )


def converter_valor_monetario(serie):
    if pd.api.types.is_numeric_dtype(serie):
        return pd.to_numeric(serie, errors="coerce")

    s = serie.astype(str).str.strip()
    s = s.str.replace("R$", "", regex=False)
    s = s.str.replace(" ", "", regex=False)

    tem_virgula = s.str.contains(",", regex=False, na=False)

    s.loc[tem_virgula] = (
        s.loc[tem_virgula]
        .str.replace(".", "", regex=False)
        .str.replace(",", ".", regex=False)
    )

    return pd.to_numeric(s, errors="coerce")


def formatar_preco_brl(valor):
    if pd.isna(valor) or valor == "":
        return ""
    try:
        return f"{float(valor):.2f}".replace(".", ",")
    except Exception:
        return ""


def carregar_excel_normal(uploaded_file, sheet_name=0):
    nome = uploaded_file.name.lower()
    uploaded_file.seek(0)

    if nome.endswith(".xlsx"):
        return pd.read_excel(uploaded_file, engine="openpyxl", sheet_name=sheet_name)

    if nome.endswith(".xls"):
        return pd.read_excel(uploaded_file, engine="xlrd", sheet_name=sheet_name)

    return pd.read_excel(uploaded_file, sheet_name=sheet_name)


def carregar_excel_bruto(uploaded_file, sheet_name=0):
    nome = uploaded_file.name.lower()
    uploaded_file.seek(0)

    if nome.endswith(".xlsx"):
        return pd.read_excel(uploaded_file, engine="openpyxl", sheet_name=sheet_name, header=None)

    if nome.endswith(".xls"):
        return pd.read_excel(uploaded_file, engine="xlrd", sheet_name=sheet_name, header=None)

    return pd.read_excel(uploaded_file, sheet_name=sheet_name, header=None)


def detectar_linha_cabecalho_cotacao(df_bruto):
    palavras = [
        "ean", "codigo ean", "código ean", "cod barra", "cód barra",
        "codigo de barras", "produto", "descricao", "descrição",
        "fabricante", "qtd", "qt", "preco", "preço", "preco un",
        "preço un", "% desc", "preco c/ desc", "preço c/ desc"
    ]

    limite = min(len(df_bruto), 25)
    melhor_linha = 0
    melhor_score = -1

    for i in range(limite):
        linha = df_bruto.iloc[i].fillna("").astype(str).tolist()
        linha_norm = [normalizar_texto(x) for x in linha]

        score = 0
        for cel in linha_norm:
            for palavra in palavras:
                if palavra in cel:
                    score += 1

        if score > melhor_score:
            melhor_score = score
            melhor_linha = i

    return melhor_linha


def detectar_linha_cabecalho_base(df_bruto):
    palavras = [
        "codigo ean", "código ean", "descricao", "descrição",
        "laboratorio", "laboratório", "st", "preco nf", "preço nf", "estoque"
    ]

    limite = min(len(df_bruto), 15)
    melhor_linha = 0
    melhor_score = -1

    for i in range(limite):
        linha = df_bruto.iloc[i].fillna("").astype(str).tolist()
        linha_norm = [normalizar_texto(x) for x in linha]

        score = 0
        for cel in linha_norm:
            for palavra in palavras:
                if palavra in cel:
                    score += 1

        if score > melhor_score:
            melhor_score = score
            melhor_linha = i

    return melhor_linha


def construir_dataframe_com_cabecalho(df_bruto, header_row):
    cab = df_bruto.iloc[header_row].fillna("").astype(str).tolist()
    dados = df_bruto.iloc[header_row + 1:].copy().reset_index(drop=True)
    dados.columns = cab
    return dados


def encontrar_coluna_por_nomes(colunas, nomes_alvo):
    mapa = {normalizar_texto(c): c for c in colunas}

    for alvo in nomes_alvo:
        alvo_norm = normalizar_texto(alvo)

        for col_norm, col_original in mapa.items():
            if alvo_norm == col_norm:
                return col_original

        for col_norm, col_original in mapa.items():
            if alvo_norm in col_norm:
                return col_original

    return None


def sugerir_coluna_ean(df):
    return encontrar_coluna_por_nomes(df.columns, [
        "ean", "codigo ean", "código ean", "cod barra", "cód barra",
        "codigo de barras", "gtin"
    ])


def sugerir_coluna_preco_real(df):
    return encontrar_coluna_por_nomes(df.columns, [
        "preço real", "preco real", "preço final", "preco final",
        "valor final", "preço venda", "preco venda"
    ])


def sugerir_coluna_st(df):
    return encontrar_coluna_por_nomes(df.columns, [
        "st", "valor st", "substituicao tributaria", "substituição tributária"
    ])


def sugerir_coluna_preco_nf(df):
    return encontrar_coluna_por_nomes(df.columns, [
        "preço nf", "preco nf", "valor nf", "preço nota", "preco nota", "nf"
    ])


def sugerir_coluna_estoque(df):
    return encontrar_coluna_por_nomes(df.columns, [
        "estoque", "saldo", "qtd estoque", "quantidade estoque",
        "disponivel", "disponível"
    ])


def sugerir_coluna_preco_cotacao(df):
    return encontrar_coluna_por_nomes(df.columns, [
        "preço un", "preco un", "preço", "preco", "valor",
        "preço unitário", "preco unitario", "preço c/ desc", "preco c/ desc"
    ])


def listar_abas_xlsx(uploaded_file):
    uploaded_file.seek(0)
    wb = load_workbook(uploaded_file, read_only=True, data_only=False)
    nomes = wb.sheetnames
    wb.close()
    return nomes


def escrever_precos_em_xlsx_original(
    uploaded_file,
    aba_nome,
    header_row_zero_based,
    preco_col_idx_zero_based,
    precos_numericos
):
    uploaded_file.seek(0)
    wb = load_workbook(uploaded_file)
    ws = wb[aba_nome]

    for i, preco in enumerate(precos_numericos):
        linha_excel = header_row_zero_based + 2 + i
        coluna_excel = preco_col_idx_zero_based + 1
        cell = ws.cell(row=linha_excel, column=coluna_excel)

        if pd.isna(preco):
            cell.value = None
        else:
            cell.value = float(preco)
            cell.number_format = "0.00"

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    return output.getvalue()


# =========================================================
# SIDEBAR
# =========================================================
st.sidebar.header("Configurações")

estoque_minimo = st.sidebar.number_input(
    "Estoque mínimo",
    min_value=0,
    value=30,
    step=1
)

st.sidebar.markdown("---")
st.sidebar.write("Fluxo:")
st.sidebar.write("1. Subir base")
st.sidebar.write("2. Subir cotação")
st.sidebar.write("3. Ajustar cabeçalhos")
st.sidebar.write("4. Associar colunas manualmente")
st.sidebar.write("5. Processar")
st.sidebar.write("6. Baixar arquivo final")

# =========================================================
# UPLOADS
# =========================================================
col1, col2 = st.columns(2)

with col1:
    st.subheader("1. Importar base da empresa")
    base_file = st.file_uploader(
        "Envie a base do pedido eletrônico",
        type=["xlsx", "xls"],
        key="base_file"
    )

with col2:
    st.subheader("2. Importar planilha de cotação")
    cotacao_file = st.file_uploader(
        "Envie a planilha de cotação",
        type=["xlsx", "xls"],
        key="cotacao_file"
    )

base_bruto = None
base_df = None
base_header_row_detectado = None

cotacao_bruto = None
cotacao_df = None
cotacao_header_row_detectado = None
aba_escolhida = None
abas_cotacao = []

# =========================================================
# BASE
# =========================================================
if base_file:
    try:
        nome_base = base_file.name.lower()

        if nome_base.endswith(".xlsx"):
            base_bruto = carregar_excel_bruto(base_file, sheet_name=0)
        else:
            base_bruto = carregar_excel_bruto(base_file)

        base_header_row_detectado = detectar_linha_cabecalho_base(base_bruto)

        st.markdown("### 1.1 Ajuste do cabeçalho da base")
        base_header_row_manual_1based = st.number_input(
            "Linha do cabeçalho da base",
            min_value=1,
            max_value=max(1, len(base_bruto)),
            value=int(base_header_row_detectado + 1),
            step=1,
            help="Na base do pedido eletrônico, normalmente é a linha onde aparecem Código EAN, ST, Preço NF e Estoque."
        )

        base_header_row = int(base_header_row_manual_1based - 1)
        base_df = construir_dataframe_com_cabecalho(base_bruto, base_header_row)
        base_df.columns = [str(c).strip() for c in base_df.columns]

        st.markdown("### Prévia da base")
        st.dataframe(base_df.head(10), use_container_width=True)

    except Exception as e:
        st.error(f"Erro ao ler a base: {e}")

# =========================================================
# COTAÇÃO
# =========================================================
if cotacao_file:
    try:
        nome_arq = cotacao_file.name.lower()

        if nome_arq.endswith(".xlsx"):
            abas_cotacao = listar_abas_xlsx(cotacao_file)
            aba_escolhida = st.selectbox("Aba da cotação", options=abas_cotacao, index=0)
            cotacao_bruto = carregar_excel_bruto(cotacao_file, sheet_name=aba_escolhida)
        else:
            aba_escolhida = "Planilha única"
            cotacao_bruto = carregar_excel_bruto(cotacao_file)

        cotacao_header_row_detectado = detectar_linha_cabecalho_cotacao(cotacao_bruto)

    except Exception as e:
        st.error(f"Erro ao ler a cotação: {e}")

if cotacao_bruto is not None:
    st.markdown("### 3. Ajuste do cabeçalho da cotação")

    header_row_manual_1based = st.number_input(
        "Linha do cabeçalho da cotação",
        min_value=1,
        max_value=max(1, len(cotacao_bruto)),
        value=int(cotacao_header_row_detectado + 1),
        step=1,
        help="Informe a linha onde estão os nomes reais das colunas."
    )

    header_row = int(header_row_manual_1based - 1)
    cotacao_df = construir_dataframe_com_cabecalho(cotacao_bruto, header_row)

    st.markdown("### Prévia da cotação")
    st.dataframe(cotacao_df.head(10), use_container_width=True)

# =========================================================
# ASSOCIAÇÃO MANUAL
# =========================================================
if base_df is not None and cotacao_df is not None:
    st.markdown("---")
    st.subheader("4. Associação das colunas")

    col_base_ean_sug = sugerir_coluna_ean(base_df)
    col_base_preco_real_sug = sugerir_coluna_preco_real(base_df)
    col_base_st_sug = sugerir_coluna_st(base_df)
    col_base_preco_nf_sug = sugerir_coluna_preco_nf(base_df)
    col_base_estoque_sug = sugerir_coluna_estoque(base_df)

    col_cot_ean_sug = sugerir_coluna_ean(cotacao_df)
    col_cot_preco_sug = sugerir_coluna_preco_cotacao(cotacao_df)

    st.info("O sistema sugere automaticamente, mas o representante confirma manualmente antes de processar.")

    s1, s2 = st.columns(2)

    with s1:
        st.markdown("#### Base da empresa")

        opcoes_base = ["-- Selecionar --"] + list(base_df.columns)

        idx_base_ean = opcoes_base.index(col_base_ean_sug) if col_base_ean_sug in opcoes_base else 0
        col_base_ean = st.selectbox("Coluna EAN da base", opcoes_base, index=idx_base_ean)

        modo_preco = st.radio(
            "Forma de preço da base",
            options=["Usar PREÇO REAL", "Calcular ST + PREÇO NF"],
            index=0 if col_base_preco_real_sug else 1
        )

        if modo_preco == "Usar PREÇO REAL":
            idx_preco_real = opcoes_base.index(col_base_preco_real_sug) if col_base_preco_real_sug in opcoes_base else 0
            col_base_preco_real = st.selectbox("Coluna PREÇO REAL", opcoes_base, index=idx_preco_real)
            col_base_st = None
            col_base_preco_nf = None
        else:
            idx_st = opcoes_base.index(col_base_st_sug) if col_base_st_sug in opcoes_base else 0
            idx_nf = opcoes_base.index(col_base_preco_nf_sug) if col_base_preco_nf_sug in opcoes_base else 0
            col_base_st = st.selectbox("Coluna ST", opcoes_base, index=idx_st)
            col_base_preco_nf = st.selectbox("Coluna PREÇO NF", opcoes_base, index=idx_nf)
            col_base_preco_real = None

        if col_base_estoque_sug:
            st.caption(f"Coluna ESTOQUE detectada automaticamente: **{col_base_estoque_sug}**")
        else:
            st.warning("Não consegui detectar automaticamente a coluna ESTOQUE da base.")

        col_base_estoque = col_base_estoque_sug

    with s2:
        st.markdown("#### Cotação do cliente")

        opcoes_cot = ["-- Selecionar --"] + list(cotacao_df.columns)

        idx_cot_ean = opcoes_cot.index(col_cot_ean_sug) if col_cot_ean_sug in opcoes_cot else 0
        idx_cot_preco = opcoes_cot.index(col_cot_preco_sug) if col_cot_preco_sug in opcoes_cot else 0

        col_cot_ean = st.selectbox("Coluna EAN da cotação", opcoes_cot, index=idx_cot_ean)
        col_cot_preco = st.selectbox("Coluna PREÇO da cotação", opcoes_cot, index=idx_cot_preco)

    st.markdown("---")
    processar = st.button("5. Processar cotação", use_container_width=True)

    if processar:
        try:
            if col_base_ean == "-- Selecionar --":
                st.error("Selecione a coluna EAN da base.")
                st.stop()

            if not col_base_estoque:
                st.error("Não consegui detectar automaticamente a coluna ESTOQUE da base.")
                st.stop()

            if col_cot_ean == "-- Selecionar --":
                st.error("Selecione a coluna EAN da cotação.")
                st.stop()

            if col_cot_preco == "-- Selecionar --":
                st.error("Selecione a coluna PREÇO da cotação.")
                st.stop()

            if modo_preco == "Usar PREÇO REAL":
                if col_base_preco_real == "-- Selecionar --":
                    st.error("Selecione a coluna PREÇO REAL da base.")
                    st.stop()
            else:
                if col_base_st == "-- Selecionar --":
                    st.error("Selecione a coluna ST da base.")
                    st.stop()
                if col_base_preco_nf == "-- Selecionar --":
                    st.error("Selecione a coluna PREÇO NF da base.")
                    st.stop()

            base_proc = base_df.copy()
            cot_proc = cotacao_df.copy()

            base_proc[col_base_ean] = limpar_ean(base_proc[col_base_ean])
            cot_proc[col_cot_ean] = limpar_ean(cot_proc[col_cot_ean])

            base_proc[col_base_estoque] = pd.to_numeric(
                base_proc[col_base_estoque],
                errors="coerce"
            ).fillna(0)

            if modo_preco == "Usar PREÇO REAL":
                base_proc["_PRECO_FINAL_"] = converter_valor_monetario(base_proc[col_base_preco_real])
            else:
                base_proc["_ST_"] = converter_valor_monetario(base_proc[col_base_st])
                base_proc["_PRECO_NF_"] = converter_valor_monetario(base_proc[col_base_preco_nf])
                base_proc["_PRECO_FINAL_"] = (
                    base_proc["_ST_"].fillna(0) + base_proc["_PRECO_NF_"].fillna(0)
                )

            base_filtrada = base_proc[base_proc[col_base_estoque] >= estoque_minimo].copy()
            base_merge = base_filtrada[[col_base_ean, "_PRECO_FINAL_"]].drop_duplicates(subset=[col_base_ean])

            resultado = cot_proc.merge(
                base_merge,
                left_on=col_cot_ean,
                right_on=col_base_ean,
                how="left"
            )

            precos_numericos = resultado["_PRECO_FINAL_"].tolist()
            precos_preview = [formatar_preco_brl(x) for x in precos_numericos]

            total_itens = len(cot_proc)
            encontrados = int(pd.notna(resultado["_PRECO_FINAL_"]).sum())
            nao_encontrados = int(pd.isna(resultado["_PRECO_FINAL_"]).sum())

            c1, c2, c3 = st.columns(3)
            c1.metric("Itens na cotação", total_itens)
            c2.metric("Encontrados", encontrados)
            c3.metric("Não encontrados", nao_encontrados)

            preview = cot_proc.copy()
            preview[col_cot_preco] = preview[col_cot_preco].astype("object")
            preview[col_cot_preco] = precos_preview

            st.markdown("### 6. Prévia final")
            st.dataframe(preview.head(30), use_container_width=True)

            linha_cab = cotacao_bruto.iloc[header_row].fillna("").astype(str).tolist()
            preco_col_idx = None

            for idx, valor in enumerate(linha_cab):
                if str(valor).strip() == str(col_cot_preco).strip():
                    preco_col_idx = idx
                    break

            if preco_col_idx is None:
                st.error("Não consegui localizar a coluna de preço na planilha original.")
                st.stop()

            nome_arq = cotacao_file.name.lower()

            if nome_arq.endswith(".xlsx"):
                arquivo_saida = escrever_precos_em_xlsx_original(
                    uploaded_file=cotacao_file,
                    aba_nome=aba_escolhida,
                    header_row_zero_based=header_row,
                    preco_col_idx_zero_based=preco_col_idx,
                    precos_numericos=precos_numericos
                )

                st.download_button(
                    label="Baixar cotação preenchida",
                    data=arquivo_saida,
                    file_name="cotacao_preenchida.xlsx",
                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                    use_container_width=True
                )
            else:
                resultado_bruto = cotacao_bruto.copy()
                resultado_bruto[preco_col_idx] = resultado_bruto[preco_col_idx].astype("object")

                for i, preco in enumerate(precos_numericos):
                    linha_real = header_row + 1 + i
                    resultado_bruto.iat[linha_real, preco_col_idx] = float(preco) if pd.notna(preco) else None

                output = io.BytesIO()
                with pd.ExcelWriter(output, engine="openpyxl") as writer:
                    resultado_bruto.to_excel(writer, index=False, header=False, sheet_name="Cotacao_Preenchida")
                output.seek(0)

                st.download_button(
                    label="Baixar cotação preenchida",
                    data=output.getvalue(),
                    file_name="cotacao_preenchida.xlsx",
                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                    use_container_width=True
                )

                st.info("Arquivo de entrada em .xls: a saída será em .xlsx, preservando a estrutura o máximo possível.")

        except Exception as e:
            st.error(f"Ocorreu um erro ao processar: {e}")

else:
    st.markdown("---")
    st.info("Envie a base e a cotação para configurar o preenchimento.")